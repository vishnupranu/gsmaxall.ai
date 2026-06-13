"""Excel (.xlsx) and CSV ingestion → text rows for the vector index.

Each row becomes a "col: value | col: value" line so tabular data is searchable
in the Knowledge OS. xlsx is parsed with openpyxl; CSV with the stdlib.
"""
from __future__ import annotations

import csv
import io


def _row_to_text(headers: list[str], values: list) -> str:
    parts = []
    for i, val in enumerate(values):
        if val is None or val == "":
            continue
        header = headers[i] if i < len(headers) and headers[i] else f"col{i + 1}"
        parts.append(f"{header}: {val}")
    return " | ".join(parts)


def parse_csv(data: bytes) -> list[str]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    return [line for r in rows[1:] if (line := _row_to_text(headers, r))]


def parse_xlsx(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for r in rows[1:]:
            line = _row_to_text(headers, list(r))
            if line:
                lines.append(f"[{ws.title}] {line}")
    wb.close()
    return lines


def parse_tabular(filename: str, data: bytes) -> list[str]:
    """Dispatch on file extension. Returns one text line per data row."""
    name = filename.lower()
    if name.endswith(".csv"):
        return parse_csv(data)
    if name.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(data)
    raise ValueError(f"unsupported tabular file: {filename} (use .csv or .xlsx)")
